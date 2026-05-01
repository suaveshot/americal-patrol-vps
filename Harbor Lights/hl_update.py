#!/usr/bin/env python3
"""
Harbor Lights Guest Parking - Daily Update Script
Reads new Harbor Lights HOA Guest Parking and Incident Report PDFs from the
Morning Reports folder tree, extracts license plates / permit numbers /
warning & tow events, and appends them to the Excel tracker with matching
styling. Also refreshes the Summary Dashboard stats.
"""

import logging
import os
import re
import json
import shutil
import sys
import tempfile
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime
from itertools import groupby

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Logging ─────────────────────────────────────────────────────────────────
BASE          = str(Path(__file__).resolve().parent.parent)
LOG_FILE      = os.path.join(BASE, "Harbor Lights", "harbor_lights.log")

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# Echo to stdout as well (for Task Scheduler output capture)
_console = logging.StreamHandler(sys.stdout)
_console.setFormatter(logging.Formatter("%(message)s"))
logging.getLogger().addHandler(_console)

# ── Configuration ─────────────────────────────────────────────────────────────
MORNING_DIR   = os.path.join(BASE, "Americal Patrol Morning Reports")
EXCEL_FILE    = os.path.join(BASE, "Harbor Lights", "Harbor Lights Guest Parking UPDATED.xlsx")
PROCESSED_LOG = os.path.join(BASE, "Harbor Lights", "processed_pdfs.json")
TEMP_FILE     = os.path.join(tempfile.gettempdir(), "hl_temp.xlsx")


def _check_excel_lock():
    """Return True if Excel has the file open (lock file present)."""
    excel_dir  = os.path.dirname(EXCEL_FILE)
    excel_name = os.path.basename(EXCEL_FILE)
    lock_file  = os.path.join(excel_dir, "~$" + excel_name)
    return os.path.exists(lock_file)


# ── Step 1: Find new PDFs ────────────────────────────────────────────────────
def find_harbor_pdfs(root):
    """Return (basename, fullpath) for every Harbor Lights PDF found recursively."""
    results = []
    for dirpath, _, filenames in os.walk(root):
        for fname in filenames:
            if not fname.lower().endswith(".pdf"):
                continue
            fl = fname.lower()
            if "harbor" in fl and "lights" in fl:
                results.append((fname, os.path.join(dirpath, fname)))
    return results


# ── Branded format detection ────────────────────────────────────────────────
# On 2026-04-02 patrol_automation stopped forwarding Connecteam-native PDFs and
# started generating its own branded PDFs via branded_pdf.py. The new PDFs have
# "GUEST PARKING CHECK" / "INCIDENT REPORT" banners, "Date April 15, 2026"
# instead of "04/15/2026", and "Plates: ..." lines instead of "License Plate #".
BRANDED_MARKER_RE       = re.compile(r"GUEST\s+PARKING\s+CHECK|INCIDENT\s+REPORT", re.IGNORECASE)
BRANDED_DATE_RE         = re.compile(r"Date\s+([A-Z][a-z]+\s+\d{1,2},\s*\d{4})")
BRANDED_PLATES_LINE_RE  = re.compile(r"Plates:\s*(.+)")
BRANDED_PLATE_TOKEN_RE  = re.compile(r"([A-Z0-9]{4,10})(?:/([A-Za-z0-9-]+))?")


def _is_branded(full_text):
    return bool(BRANDED_MARKER_RE.search(full_text[:2000]))


def _branded_date(full_text):
    """Return MM/DD/YYYY string if the PDF has a branded 'Date Month D, YYYY' header."""
    m = BRANDED_DATE_RE.search(full_text)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%B %d, %Y").strftime("%m/%d/%Y")
    except ValueError:
        return None


# ── Step 2: Parse Incident Reports ──────────────────────────────────────────
TOW_RE = re.compile(r"\b(?:tow(?:ed|ing)?|vehicle\s+towed|was\s+towed)\b", re.IGNORECASE)


def clean_plate(raw):
    if not raw:
        return None
    plate = re.sub(r"^LP\s*#\s*", "", raw.strip(), flags=re.IGNORECASE)
    plate = re.sub(r"\s+", "", plate).upper()
    return plate if plate else None


def parse_incident_pdfs(incident_pdfs):
    towed_pairs   = set()
    warning_pairs = set()

    for bn, fpath in incident_pdfs:
        try:
            with pdfplumber.open(fpath) as pdf:
                page_texts = [(page.extract_text() or "") for page in pdf.pages]
            full_text = "\n".join(page_texts)

            # Resolve a PDF-level date once. Branded PDFs put the date in the
            # page-1 meta table only; later pages (photos, ticket scans) have no
            # date and would otherwise be skipped.
            old_m = re.search(r"(\d{2}/\d{2}/\d{4})", full_text)
            pdf_date = old_m.group(1) if old_m else _branded_date(full_text)

            for text in page_texts:
                if re.match(r"^\s*\d+/\d+\s*$", text.strip()):
                    continue
                page_m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
                date_str = page_m.group(1) if page_m else pdf_date
                if not date_str:
                    continue
                is_towed = bool(TOW_RE.search(text))
                plates = []
                for m in re.finditer(
                    r"LP\s*#\s*([A-Z0-9][A-Z0-9 ]{0,12}[A-Z0-9])(?=\W|\Z)",
                    text, re.IGNORECASE
                ):
                    p = clean_plate(m.group(1))
                    if p:
                        plates.append(p)
                for m in re.finditer(
                    r"license\s+plate\s+([A-Z0-9]{5,8})\b", text, re.IGNORECASE
                ):
                    p = clean_plate(m.group(1))
                    if p:
                        plates.append(p)
                for plate in set(plates):
                    if is_towed:
                        towed_pairs.add((date_str, plate))
                    else:
                        warning_pairs.add((date_str, plate))
        except Exception as e:
            log.warning(f"Could not parse incident PDF {bn}: {e}")

    warning_pairs -= towed_pairs   # towed takes priority
    return towed_pairs, warning_pairs


# ── Step 3: Parse Guest Parking PDFs ─────────────────────────────────────────
def clean_permit(raw):
    if raw is None:
        return None
    r = raw.upper().strip()
    if "HANDICAP" in r or "HAND" in r:
        return "HANDICAP"
    if "NO PERMIT" in r or "NOPERM" in r:
        return "NO PERMIT"
    if r in ("N/A", "N-A", "NA"):
        return None
    digits = re.sub(r"[^0-9]", "", r)
    if digits:
        return int(digits)
    return None


LP_HEADER_RE = re.compile(
    r"License\s+Plate\s*[#&]?\s*"
    r"(?:(?:Parking\s+)?(?:Pass|Permit)(?:\s+Number)?\s*[:#]?|"
    r"&\s*Parking\s+Permit(?:\s+Number)?\s*[:#]?)?",
    re.IGNORECASE,
)
BLOCK_END_RE = re.compile(
    r"^\s*(?:pictures?|\d+/\d+|harbor\s+lights|license\s+plate"
    r"|[\w\s]+guest\s+parking|\s*)$",
    re.IGNORECASE,
)


def flush_lp_block(lp_lines):
    """Parse accumulated LP text lines; return [(plate, permit), ...]."""
    results, seen = [], set()
    raw = " ".join(lp_lines)
    raw = re.sub(r"\.\s*\(", " (", raw)
    raw = re.sub(r"LP\s*#\s*", "LP#", raw, flags=re.IGNORECASE)

    for m in re.finditer(r"LP#([A-Z0-9\s]{2,20}?)\s*\(+([^)]+)\)", raw, re.IGNORECASE):
        plate  = clean_plate(m.group(1))
        permit = clean_permit(m.group(2))
        if plate and plate not in seen:
            results.append((plate, permit))
            seen.add(plate)

    for m in re.finditer(r"LP#([A-Z0-9\s]{2,20}?)(?=\s*(?:LP#|\Z))", raw, re.IGNORECASE):
        plate = clean_plate(m.group(1))
        if plate and plate not in seen:
            results.append((plate, None))
            seen.add(plate)

    # New format: PLATE/PERMIT (e.g. "6WAE699/082", "9JFT175/handicap", "7NRR195/N-A")
    for m in re.finditer(r"\b([A-Z0-9]{4,10})/([A-Za-z0-9-]+)", raw):
        plate  = clean_plate(m.group(1))
        permit = clean_permit(m.group(2))
        if not plate:
            continue
        if plate in seen:
            for idx, existing in enumerate(results):
                if existing[0] == plate and existing[1] is None and permit is not None:
                    results[idx] = (plate, permit)
            continue
        results.append((plate, permit))
        seen.add(plate)

    for m in re.finditer(r"\b([A-Z0-9]{5,8})\b", raw, re.IGNORECASE):
        token = m.group(1).upper()
        if (token not in seen
                and not token.startswith("LP")
                and any(c.isdigit() for c in token)
                and any(c.isalpha() for c in token)):
            results.append((token, None))
            seen.add(token)

    for m in re.finditer(r"([A-Z0-9]{5,8})\s*\(+([^)]+)\)", raw, re.IGNORECASE):
        plate  = m.group(1).upper()
        permit = clean_permit(m.group(2))
        if any(c.isdigit() for c in plate) and any(c.isalpha() for c in plate):
            existing = next((i for i, r in enumerate(results) if r[0] == plate), None)
            if existing is not None:
                results[existing] = (plate, permit)
            elif plate not in seen:
                results.append((plate, permit))
                seen.add(plate)
    return results


def _parse_branded_guest_text(full_text):
    """Extract (date_str, [(plate, permit), ...]) from a branded GUEST PARKING CHECK PDF.

    New format (post-2026-04-02 patrol_automation cutover): date appears once
    in the meta table ("Date April 15, 2026") and plate lists are prefixed
    with "Plates: ...". Officers enter plates in several shapes on the same
    day — `LP#PLATE(PERMIT)`, `PLATE/PERMIT`, or bare plates — so delegate
    the actual plate+permit parsing to flush_lp_block(), which already
    handles every format the old PDFs used and filters out non-plate words.
    """
    date_str = _branded_date(full_text)
    if not date_str:
        return None, []
    plate_text_lines = BRANDED_PLATES_LINE_RE.findall(full_text)
    if not plate_text_lines:
        return date_str, []
    return date_str, flush_lp_block(plate_text_lines)


def parse_guest_pdfs(guest_pdfs):
    parking_data = {}

    for bn, fpath in guest_pdfs:
        try:
            with pdfplumber.open(fpath) as pdf:
                full_text = ""
                for page in pdf.pages:
                    full_text += (page.extract_text() or "") + "\n"

            # New branded format (post-2026-04-02). Keep the legacy path below
            # for any pre-cutover PDFs or manual re-parses.
            if _is_branded(full_text):
                date_str, entries = _parse_branded_guest_text(full_text)
                if date_str and entries:
                    parking_data.setdefault(date_str, []).extend(entries)
                elif date_str:
                    log.info(f"Branded guest parking PDF {bn}: no plates on {date_str} (no vehicles found)")
                else:
                    log.warning(f"Branded guest parking PDF {bn}: could not resolve date")
                continue

            current_date = None
            sections = re.split(r"Harbor Lights HOA Guest Parking\n", full_text)
            for section in sections:
                if not section.strip():
                    continue
                date_m = re.search(r"(\d{2}/\d{2}/\d{4})", section)
                if date_m:
                    current_date = date_m.group(1)
                if not current_date:
                    continue

                lines, in_lp_block, lp_lines, entries = section.splitlines(), False, [], []
                for line in lines:
                    if LP_HEADER_RE.search(line):
                        if lp_lines:
                            entries.extend(flush_lp_block(lp_lines))
                            lp_lines = []
                        in_lp_block = True
                        data_part = LP_HEADER_RE.sub("", line).strip()
                        if data_part:
                            lp_lines.append(data_part)
                    elif in_lp_block:
                        if BLOCK_END_RE.match(line):
                            entries.extend(flush_lp_block(lp_lines))
                            lp_lines, in_lp_block = [], False
                        else:
                            stripped = line.strip()
                            if stripped:
                                lp_lines.append(stripped)
                if lp_lines:
                    entries.extend(flush_lp_block(lp_lines))

                for plate, permit in entries:
                    parking_data.setdefault(current_date, []).append((plate, permit))

        except Exception as e:
            log.warning(f"Could not parse guest parking PDF {bn}: {e}")

    return parking_data


# ── Step 4–6: Excel update ────────────────────────────────────────────────────
THIN        = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center")

STATUS_STYLE = {
    "TOWED":     ("C00000", "FFFFFF", True),
    "WARNING":   ("FF6600", "FFFFFF", True),
    "HANDICAP":  ("2E75B6", "FFFFFF", True),
    "NO PERMIT": ("FFD966", "000000", False),
}


def _font(bold=False, color="000000"):
    return Font(name="Aptos Narrow", size=10, bold=bold, color=color)


def _fill(hex6):
    return PatternFill(fill_type="solid", fgColor="FF" + hex6)


def write_data_row(ws, row_num, date_val, plate, permit, row_fill_hex):
    ca = ws.cell(row=row_num, column=1, value=date_val)
    ca.font, ca.fill = _font(bold=(date_val is not None)), _fill(row_fill_hex)
    ca.alignment, ca.border = CENTER, THIN_BORDER
    if date_val is not None:
        ca.number_format = "d-mmm-yy"

    cb = ws.cell(row=row_num, column=2, value=plate)
    cb.font, cb.fill = _font(), _fill(row_fill_hex)
    cb.alignment, cb.border = CENTER, THIN_BORDER

    permit_val = permit
    if isinstance(permit, str) and permit.upper() not in STATUS_STYLE:
        digits = re.sub(r"[^0-9]", "", permit)
        permit_val = int(digits) if digits else (permit if permit.strip() else None)
    cc = ws.cell(row=row_num, column=3, value=permit_val)
    sk = str(permit_val).upper() if permit_val is not None else ""
    if sk in STATUS_STYLE:
        bg, fg, bold = STATUS_STYLE[sk]
        cc.font, cc.fill = _font(bold=bold, color=fg), _fill(bg)
    else:
        cc.font, cc.fill = _font(), _fill(row_fill_hex)
    cc.alignment, cc.border = CENTER, THIN_BORDER


def update_excel(all_rows, parking_data, towed_pairs, warning_pairs):
    """Append rows to Excel and refresh Summary Dashboard. Returns True on success."""
    # Check for Excel lock before attempting to open
    if _check_excel_lock():
        raise OSError(
            f"Excel file is open in another process. "
            f"Close '{os.path.basename(EXCEL_FILE)}' in Excel and re-run."
        )

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Determine target sheet — use year sheet if present (new format), else legacy sheet
    current_year = str(datetime.now().year)
    if current_year in wb.sheetnames:
        ws = wb[current_year]
    elif "Harbor Lights HOA" in wb.sheetnames:
        ws = wb["Harbor Lights HOA"]
    else:
        # Create a new year sheet if neither exists
        ws = wb.create_sheet(current_year)
        log.warning(f"Created new year sheet '{current_year}' in Excel.")

    last_fill = None
    for r in range(ws.max_row, 0, -1):
        cell = ws.cell(row=r, column=1)
        if cell.value is not None:
            last_fill = cell.fill.fgColor.rgb[-6:].upper()
            break

    current_fill      = "FFFFFF" if last_fill == "EEF2F7" else "EEF2F7"
    current_write_row = ws.max_row + 2

    first_group = True
    for date_obj, group_iter in groupby(all_rows, key=lambda x: x[0]):
        entries = list(group_iter)
        if not first_group:
            current_write_row += 1
            current_fill = "FFFFFF" if current_fill == "EEF2F7" else "EEF2F7"
        first_entry = True
        for _, plate, permit in entries:
            write_data_row(ws, current_write_row,
                           date_obj if first_entry else None,
                           plate, permit, current_fill)
            first_entry, current_write_row = False, current_write_row + 1
        first_group = False

    # Refresh Summary Dashboard
    _refresh_summary(wb, ws)

    wb.save(TEMP_FILE)
    shutil.copy2(TEMP_FILE, EXCEL_FILE)
    log.info(f"Excel saved -> {EXCEL_FILE}")
    return True


def _refresh_dashboard_new(wb):
    """Refresh KPIs and top-10 table in the new-format Dashboard sheet."""
    dash = wb["Dashboard"]

    # Collect all rows from all year sheets
    all_rows_data = []
    for sheet_name in wb.sheetnames:
        if sheet_name.isdigit():
            ws_year = wb[sheet_name]
            current_date = None
            for row in ws_year.iter_rows(min_row=2, values_only=True):
                date_val, plate, status = row[0], row[1], row[2]
                if date_val is not None:
                    current_date = date_val if isinstance(date_val, datetime) else None
                if plate and current_date:
                    all_rows_data.append((current_date, str(plate), str(status) if status else ""))

    if not all_rows_data:
        return

    total     = len(all_rows_data)
    unique    = len(set(r[1] for r in all_rows_data))
    days      = len(set(r[0].date() for r in all_rows_data if isinstance(r[0], datetime)))
    daily_avg = round(total / days, 1) if days else 0
    towed     = sum(1 for r in all_rows_data if r[2].upper() == "TOWED")
    warnings  = sum(1 for r in all_rows_data if r[2].upper() == "WARNING")

    # KPI values are in row 7, cols B(2)–G(7)
    kpi_vals = [total, unique, days, daily_avg, towed, warnings]
    for col, val in enumerate(kpi_vals, start=2):
        dash.cell(row=7, column=col).value = val

    # Update generated timestamp (row 3 title banner)
    dash["B3"].value = f"Generated {datetime.now().strftime('%B %d, %Y')}"

    # Update top-10 table: scan from row 1 to find "TOP 10" section header
    plate_counter   = Counter(r[1] for r in all_rows_data)
    plate_last_seen = {}
    for date_obj, plate, _ in all_rows_data:
        if isinstance(date_obj, datetime):
            if plate not in plate_last_seen or date_obj > plate_last_seen[plate]:
                plate_last_seen[plate] = date_obj

    # Find top-10 header row
    top10_data_row = None
    for r in range(1, dash.max_row + 1):
        val = str(dash.cell(row=r, column=2).value or "")
        if "TOP 10" in val.upper() or "RANK" in val.upper():
            top10_data_row = r + 1
            break

    if top10_data_row:
        for i, (plate, cnt) in enumerate(plate_counter.most_common(10)):
            r = top10_data_row + i + 1  # +1 for the header row within section
            dash.cell(row=r, column=2).value = i + 1
            dash.cell(row=r, column=3).value = plate
            dash.cell(row=r, column=4).value = cnt
            last = plate_last_seen.get(plate)
            dash.cell(row=r, column=5).value = last.strftime("%m/%d/%Y") if last else ""


def _refresh_summary(wb, ws):
    # Support both new format (Dashboard) and legacy (Summary Dashboard)
    if "Summary Dashboard" in wb.sheetnames:
        ws2 = wb["Summary Dashboard"]
    elif "Dashboard" in wb.sheetnames:
        # New format: update KPI row and top-10 table in Dashboard sheet
        _refresh_dashboard_new(wb)
        return
    else:
        log.warning("No dashboard sheet found — skipping summary refresh.")
        return

    total_vehicles, incidents = 0, 0
    unique_plates, days_covered = set(), set()
    monthly = defaultdict(lambda: {
        "total": 0, "permit": 0, "handicap": 0,
        "warning": 0, "no_permit": 0, "towed": 0, "other": 0
    })
    permit_cats     = Counter()
    plate_counter   = Counter()
    plate_last_seen = {}
    current_date    = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val, plate, permit = row[0], row[1], row[2]
        if date_val is not None and isinstance(date_val, datetime):
            current_date = date_val
        if plate is None or current_date is None:
            continue

        total_vehicles += 1
        plate_str = str(plate).upper().strip()
        unique_plates.add(plate_str)
        days_covered.add(current_date.date())
        plate_counter[plate_str] += 1
        if plate_str not in plate_last_seen or current_date > plate_last_seen[plate_str]:
            plate_last_seen[plate_str] = current_date

        mkey  = current_date.strftime("%b %Y")
        monthly[mkey]["total"] += 1
        p_str = str(permit).upper().strip() if permit is not None else ""

        if p_str == "WARNING":
            monthly[mkey]["warning"]  += 1; permit_cats["warning"]  += 1; incidents += 1
        elif p_str == "TOWED":
            monthly[mkey]["towed"]    += 1; permit_cats["towed"]    += 1; incidents += 1
        elif p_str == "HANDICAP":
            monthly[mkey]["handicap"] += 1; permit_cats["handicap"] += 1
        elif p_str == "NO PERMIT":
            monthly[mkey]["no_permit"]+= 1; permit_cats["no_permit"]+= 1
        elif permit is not None and p_str not in ("", "NONE"):
            monthly[mkey]["permit"]   += 1; permit_cats["permit"]   += 1
        else:
            monthly[mkey]["other"]    += 1; permit_cats["other"]    += 1

    ws2["B7"] = total_vehicles
    ws2["C7"] = len(unique_plates)
    ws2["E7"] = len(days_covered)
    ws2["G7"] = incidents

    month_row_map, total_row = {}, None
    for r in range(12, ws2.max_row + 1):
        val = ws2.cell(row=r, column=2).value
        if val == "TOTAL":
            total_row = r; break
        if isinstance(val, str) and val.strip():
            month_row_map[val.strip()] = r

    months_with_data = sorted(
        [m for m in monthly if monthly[m]["total"] > 0],
        key=lambda m: datetime.strptime(m, "%b %Y"),
    )
    for mkey in months_with_data:
        d = monthly[mkey]
        if mkey in month_row_map:
            r = month_row_map[mkey]
        else:
            insert_at = total_row if total_row else ws2.max_row + 1
            ws2.insert_rows(insert_at)
            r = insert_at
            if total_row: total_row += 1
            ws2.cell(row=r, column=2).value = mkey
            month_row_map[mkey] = r
        for col, key in enumerate(["total","permit","handicap","warning","no_permit","towed"], start=3):
            ws2.cell(row=r, column=col).value = d[key]

    if total_row:
        ws2.cell(row=total_row, column=3).value = total_vehicles
        for col, key in enumerate(["permit","handicap","warning","no_permit","towed"], start=4):
            ws2.cell(row=total_row, column=col).value = permit_cats[key]

    permit_total = total_vehicles or 1
    for r in range(43, ws2.max_row + 1):
        label = ws2.cell(row=r, column=2).value
        if not isinstance(label, str): continue
        lbl, count = label.strip().lower(), None
        if "valid"      in lbl: count = permit_cats["permit"]
        elif "handicap" in lbl: count = permit_cats["handicap"]
        elif "warning"  in lbl: count = permit_cats["warning"]
        elif "no permit"in lbl: count = permit_cats["no_permit"]
        elif "towed"    in lbl: count = permit_cats["towed"]
        elif "other"    in lbl or "blank" in lbl: count = permit_cats["other"]
        if count is not None:
            ws2.cell(row=r, column=3).value = count
            ws2.cell(row=r, column=4).value = count / permit_total

    for i, (plate, cnt) in enumerate(plate_counter.most_common(10)):
        r = 66 + i
        ws2.cell(row=r, column=2).value = i + 1
        ws2.cell(row=r, column=3).value = plate
        ws2.cell(row=r, column=4).value = cnt
        ws2.cell(row=r, column=5).value = plate_last_seen.get(plate)

    for r in range(ws2.max_row, max(1, ws2.max_row - 5), -1):
        cell = ws2.cell(row=r, column=2)
        if isinstance(cell.value, str) and "Generated" in cell.value:
            cell.value = (
                f"Generated {datetime.now().strftime('%B %d, %Y')}"
                "  |  Americal Patrol Security Services  |  Harbor Lights HOA"
            )
            break


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("Harbor Lights update starting")

    # Report health at end of run (imported here to avoid top-level import failure
    # if shared_utils isn't on sys.path when running standalone)
    try:
        sys.path.insert(0, BASE)
        from shared_utils.health_reporter import report_status
    except ImportError:
        report_status = None

    try:
        # Load processed log
        if os.path.exists(PROCESSED_LOG):
            with open(PROCESSED_LOG) as fp:
                processed = set(json.load(fp))
        else:
            processed = set()

        all_pdfs      = find_harbor_pdfs(MORNING_DIR)
        new_pdfs      = [(bn, fp) for bn, fp in all_pdfs if bn not in processed]
        guest_pdfs    = [(bn, fp) for bn, fp in new_pdfs
                         if "guest_parking" in bn.lower() or "guest parking" in bn.lower()]
        incident_pdfs = [(bn, fp) for bn, fp in new_pdfs
                         if "incident_report" in bn.lower() or "incident report" in bn.lower()]

        log.info(f"New Guest Parking PDFs:   {[bn for bn, _ in guest_pdfs]}")
        log.info(f"New Incident Report PDFs: {[bn for bn, _ in incident_pdfs]}")

        if not new_pdfs:
            log.info("No new PDFs found. Nothing to update.")
            if report_status:
                report_status("harbor_lights", "ok", "No new PDFs to process")
            return

        towed_pairs, warning_pairs = parse_incident_pdfs(incident_pdfs)
        parking_data               = parse_guest_pdfs(guest_pdfs)

        # Silent-failure guard: if we processed guest PDFs but extracted zero
        # plates across all of them, the PDF format has likely drifted. The
        # 2026-04-02 branded-PDF cutover hid exactly this bug for 14 days.
        if guest_pdfs and not parking_data:
            err_msg = (
                f"Guest parking PDFs processed but ZERO plates extracted — "
                f"likely PDF format change. PDFs: {[bn for bn, _ in guest_pdfs]}"
            )
            log.error(err_msg)
            if report_status:
                report_status(
                    "harbor_lights", "error",
                    f"{len(guest_pdfs)} guest parking PDF(s) yielded 0 plates — "
                    "format may have changed",
                )

        log.info(f"Towed: {towed_pairs}")
        log.info(f"Warning: {warning_pairs}")
        log.info("Guest parking data:")
        for d, entries in sorted(parking_data.items()):
            log.info(f"  {d}: {entries}")

        # Build rows
        all_rows  = []
        all_dates = sorted(set(
            list(parking_data.keys())
            + [d for d, _ in towed_pairs]
            + [d for d, _ in warning_pairs]
        ))

        for date_str in all_dates:
            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
            except ValueError:
                continue
            entries, seen_plates = parking_data.get(date_str, []), set()
            for plate, permit in entries:
                if (date_str, plate) in towed_pairs:
                    permit = "TOWED"
                elif (date_str, plate) in warning_pairs:
                    permit = "WARNING"
                all_rows.append((date_obj, plate, permit))
                seen_plates.add(plate)
            for d, plate in towed_pairs:
                if d == date_str and plate not in seen_plates:
                    all_rows.append((date_obj, plate, "TOWED"))
                    seen_plates.add(plate)
            for d, plate in warning_pairs:
                if d == date_str and plate not in seen_plates:
                    all_rows.append((date_obj, plate, "WARNING"))
                    seen_plates.add(plate)

        all_rows.sort(key=lambda x: x[0])
        towed_count   = sum(1 for r in all_rows if r[2] == "TOWED")
        warning_count = sum(1 for r in all_rows if r[2] == "WARNING")
        log.info(f"Rows to insert: {len(all_rows)}  (TOWED={towed_count}, WARNING={warning_count})")

        # Write Excel
        update_excel(all_rows, parking_data, towed_pairs, warning_pairs)

        # Update processed log
        newly_processed = sorted(set(list(processed) + [bn for bn, _ in (guest_pdfs + incident_pdfs)]))
        with open(PROCESSED_LOG, "w") as fp:
            json.dump(newly_processed, fp, indent=2)

        log.info(f"Harbor Lights Update complete — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        log.info(f"  Guest PDFs processed  : {len(guest_pdfs)}")
        log.info(f"  Incident PDFs processed: {len(incident_pdfs)}")
        log.info(f"  Rows added to Excel   : {len(all_rows)}")
        log.info(f"    Towed flags         : {towed_count}")
        log.info(f"    Warning flags       : {warning_count}")
        log.info("  Dashboard refreshed   : Yes")

        if report_status:
            report_status(
                "harbor_lights", "ok",
                f"{len(all_rows)} rows added (towed={towed_count}, warning={warning_count})",
                metrics={"rows_added": len(all_rows), "towed": towed_count, "warnings": warning_count},
            )

    except OSError as e:
        # Excel lock or file permission error
        log.error(f"ERROR: File access failed: {e}")
        if report_status:
            report_status("harbor_lights", "error", str(e))
        sys.exit(1)

    except Exception as e:
        log.error(f"ERROR: Unexpected failure: {e}", exc_info=True)
        if report_status:
            report_status("harbor_lights", "error", str(e))
        sys.exit(1)


if __name__ == "__main__":
    main()
