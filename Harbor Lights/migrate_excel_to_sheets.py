"""
One-shot migration: Harbor Lights Guest Parking UPDATED.xlsx → Google Sheet.

Reads every year-tab and the Dashboard tab from the legacy Excel file and
writes them into the Sheet identified by HL_SHEET_ID. After this runs once
successfully, hl_update.py and parking_audit.py read/write Sheets only and
the Excel file becomes a frozen archive.

Usage (on Sam's PC, with .env loaded):

    cd "Harbor Lights"
    # 1. create a fresh blank Google Sheet, copy its ID from the URL,
    #    set HL_SHEET_ID in Americal Patrol/.env
    # 2. run:
    python migrate_excel_to_sheets.py

Idempotent: re-running clears the Sheet's year tabs and re-imports from Excel.
Safe to run repeatedly during testing.
"""

from __future__ import annotations

import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
from sheets_client import (  # noqa: E402
    open_sheet,
    get_or_create_year_sheet,
    append_data_rows,
    refresh_dashboard,
    YEAR_HEADER,
)

EXCEL_FILE = SCRIPT_DIR / "Harbor Lights Guest Parking UPDATED.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


def _read_year_sheet(ws):
    """Yield (date, plate, status) tuples for every data row in an openpyxl
    year worksheet. Inherits the date from the previous non-blank row (matches
    the original Excel's grouped-row pattern)."""
    current_date = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[0] if len(row) > 0 else None
        plate    = row[1] if len(row) > 1 else None
        status   = row[2] if len(row) > 2 else None

        if isinstance(date_val, datetime):
            current_date = date_val
        if plate and current_date:
            status_str = "" if status is None else str(status)
            yield (current_date, str(plate).strip(), status_str)


def main() -> None:
    load_dotenv(SCRIPT_DIR.parent / ".env")
    if not os.environ.get("HL_SHEET_ID"):
        raise SystemExit("Set HL_SHEET_ID in .env (the new blank Google Sheet's ID).")
    if not EXCEL_FILE.exists():
        raise SystemExit(f"Excel file not found at {EXCEL_FILE}.")

    log.info(f"Reading {EXCEL_FILE}")
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)

    spreadsheet = open_sheet()
    log.info(f"Opened target Sheet: {spreadsheet.title}")

    # Clear out any pre-existing year tabs to keep the migration idempotent
    # (don't touch Dashboard / Sheet1 — refresh_dashboard rebuilds Dashboard
    # at the end).
    for ws in spreadsheet.worksheets():
        if ws.title.isdigit():
            log.info(f"Clearing existing year tab '{ws.title}'")
            spreadsheet.del_worksheet(ws)

    total_rows = 0
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit() and sheet_name != "Harbor Lights HOA":
            continue
        target_year = sheet_name if sheet_name.isdigit() else "2024"  # legacy tab → 2024 bucket
        rows = list(_read_year_sheet(wb[sheet_name]))
        if not rows:
            log.info(f"Source tab '{sheet_name}' is empty, skipping.")
            continue

        ws = get_or_create_year_sheet(spreadsheet, target_year)

        # Write header on top so the migration looks identical to a fresh setup.
        ws.update("A1:C1", [YEAR_HEADER], value_input_option="USER_ENTERED")

        # Mark continuation rows by stripping date when same-as-previous (matches
        # the visual pattern of the source Excel).
        flagged = []
        last_d = None
        for d, plate, status in rows:
            flagged.append((d if d != last_d else None, plate, status))
            last_d = d

        log.info(f"Migrating {len(flagged)} rows from '{sheet_name}' → year tab '{target_year}'")
        append_data_rows(ws, flagged)
        total_rows += len(flagged)

    log.info(f"Wrote {total_rows} rows total. Refreshing Dashboard …")
    refresh_dashboard(spreadsheet)
    log.info("Migration complete. Open the Sheet and spot-check before disabling Excel reads.")


if __name__ == "__main__":
    main()
