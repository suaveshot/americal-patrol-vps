#!/usr/bin/env python3
"""
Harbor Lights Guest Parking – Rule D Audit
-------------------------------------------
Reads all vehicle records from the Excel tracker, checks for Rule D violations,
and creates a Gmail draft addressed to the Harbor Lights HOA group.

Rule D: A "Guest" may park for no more than:
  • 4 consecutive days, OR
  • 10 non-consecutive days within a 30-day period

Run daily at 7:30 AM via Windows Task Scheduler.
"""

import os
import base64
import logging
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from collections import defaultdict

import openpyxl
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE        = Path(__file__).resolve().parent.parent
EXCEL_FILE  = BASE / "Harbor Lights" / "Harbor Lights Guest Parking UPDATED.xlsx"
PATROL_AUTH = BASE / "patrol_automation"
TOKEN_PATH  = PATROL_AUTH / "token.json"
LOG_PATH    = BASE / "Harbor Lights" / "parking_audit.log"

# ── Recipients ─────────────────────────────────────────────────────────────────
HOA_RECIPIENTS = [
    "wendellthompsonrealty@gmail.com",
    "smork@pmpmanage.com",
    "beautyandbighead@outlook.com",
    "lisaaguailar@gmail.com",
    "marilynmatthews@me.com",
    "angieped418@gmail.com",
    "breanna@fischercpa.com",
    "gj@gjlunlimited.com",
]

CC_LIST = [
    "salarcon@americalpatrol.com",
    "don@americalpatrol.com",
]

# ── Rule D thresholds ─────────────────────────────────────────────────────────
MAX_CONSECUTIVE      = 4   # more than this = violation
MAX_DAYS_IN_30_DAY   = 10  # more than this = violation

# ── Exclusions (never flagged for Rule D) ─────────────────────────────────────
# Handicap permits are excluded (long-term medical accommodations).
# 9LFC906 is excluded per HOA management direction.
EXCLUDED_PLATES  = {"9LFC906"}
EXCLUDED_PERMITS = {"HANDICAP"}

# ── Gmail scopes ──────────────────────────────────────────────────────────────
SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.compose",
    "https://www.googleapis.com/auth/gmail.send",
]

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    filename=str(LOG_PATH),
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ── Gmail auth ────────────────────────────────────────────────────────────────
def _client_config():
    return {
        "installed": {
            "client_id":     os.environ["GOOGLE_CLIENT_ID"],
            "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
            "project_id":    "americal-patrol-automation",
            "auth_uri":      "https://accounts.google.com/o/oauth2/auth",
            "token_uri":     "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "redirect_uris": ["http://localhost"],
        }
    }


def get_gmail_service():
    creds = None
    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_config(_client_config(), SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_PATH.write_text(creds.to_json())
    return build("gmail", "v1", credentials=creds)


# ── Read Excel ────────────────────────────────────────────────────────────────
def read_parking_records():
    """
    Returns list of (date, plate_str, permit_str_or_None, status_str).
    Handles the date pattern: date appears on the first row of each group;
    subsequent rows in the same group have a None date cell.

    status_str is the raw Permit/Status column value (e.g. "82", "WARNING",
    "TOWED", "HANDICAP", "NO PERMIT", "Guest", or None).
    """
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Support new year-based sheets ("2026") and legacy sheet name
    current_year = str(date.today().year)
    if current_year in wb.sheetnames:
        ws = wb[current_year]
    elif "Harbor Lights HOA" in wb.sheetnames:
        ws = wb["Harbor Lights HOA"]
    else:
        # Fall back to first non-Dashboard sheet
        ws = next(
            (wb[s] for s in wb.sheetnames if s not in ("Dashboard", "Sheet1")),
            wb.active,
        )

    records = []
    current_date = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[0]
        plate    = row[1]
        status   = row[2] if len(row) > 2 else None

        # Advance current date when a new date cell appears
        if date_val is not None and isinstance(date_val, datetime):
            current_date = date_val.date()

        if plate is None or current_date is None:
            continue

        plate_str = str(plate).strip().upper()
        if not plate_str:
            continue

        status_str = str(status).strip() if status is not None else None

        # For permit tracking, extract a clean permit value from the status field.
        # Numeric statuses are permit numbers; keyword statuses are flags.
        permit_str = status_str

        records.append((current_date, plate_str, permit_str, status_str))

    log.info("Loaded %d parking records from Excel.", len(records))
    return records


# ── Citation thresholds ────────────────────────────────────────────────────────
# A "citation" is any WARNING entry for the plate within the 30-day window.
#
# Permitted vehicles (Rule D violators):
#   After 2 citations → recommend tow on next offense.
#   After 2+ citations and still appearing → recommend immediate tow.
CITATION_TOW_THRESHOLD = 2
#
# Unpermitted vehicles (no valid parking pass):
#   2 WARNINGs in 30 days → notify HOA for tow authorization on 3rd offense.
UNPERMITTED_TOW_CITATIONS = 2


def _is_valid_permit(status_str):
    """Return True if status_str represents a valid numeric parking permit."""
    if status_str is None:
        return False
    s = str(status_str).strip()
    return s.isdigit()


# ── Rule D violation check ────────────────────────────────────────────────────
def find_violations(records, today=None):
    """
    Returns (rule_d_violations, citation_violations).

    rule_d_violations: dict  plate -> {
        'permit', 'total_days', 'first_seen', 'last_seen',
        'violations': [{'type': 'consecutive'|'30-day', 'details': str}, ...],
        'citation_count', 'citation_dates',
    }

    citation_violations: dict  plate -> {
        'citation_count', 'citation_dates', 'first_seen', 'last_seen',
    }
    Unpermitted vehicles with >= UNPERMITTED_TOW_CITATIONS WARNINGs in 30 days.

    today defaults to date.today() if not provided.
    """
    if today is None:
        today = date.today()
    window_start = today - timedelta(days=29)   # rolling 30-day window: [today-29 … today]

    # Collect unique dates per plate; track most recent non-null permit; count citations
    plate_dates       = defaultdict(set)
    plate_permit      = {}   # plate -> (latest_date, permit_str)
    plate_citations   = defaultdict(list)   # plate -> [citation_date, ...]
    plate_has_permit  = set()  # plates that have had a valid numeric permit

    for record in records:
        rec_date, plate, permit, status = record
        plate_dates[plate].add(rec_date)
        if permit is not None and permit not in ("None", "", "Guest"):
            prev = plate_permit.get(plate)
            if prev is None or rec_date >= prev[0]:
                plate_permit[plate] = (rec_date, permit)
        if _is_valid_permit(status):
            plate_has_permit.add(plate)
        # Count WARNINGs within the 30-day window as citations
        if (status is not None
                and str(status).upper() == "WARNING"
                and window_start <= rec_date <= today):
            plate_citations[plate].append(rec_date)

    results = {}

    for plate, dates_set in plate_dates.items():
        # Skip excluded plates and any plate whose most recent permit is excluded
        if plate in EXCLUDED_PLATES:
            continue
        permit_info = plate_permit.get(plate)
        if permit_info and str(permit_info[1]).upper() in EXCLUDED_PERMITS:
            continue
        # Only consider dates within the rolling 30-day window for both checks
        sorted_dates = sorted(d for d in dates_set if window_start <= d <= today)
        if not sorted_dates:
            continue
        plate_violations = []

        # ── Check 1: consecutive days ──────────────────────────────────────
        # Walk sorted dates; find any run longer than MAX_CONSECUTIVE.
        run_start = sorted_dates[0]
        run_len   = 1
        worst_run = (run_start, 1)

        for i in range(1, len(sorted_dates)):
            gap = (sorted_dates[i] - sorted_dates[i - 1]).days
            if gap == 1:
                run_len += 1
                if run_len > worst_run[1]:
                    worst_run = (run_start, run_len)
            else:
                run_start = sorted_dates[i]
                run_len   = 1

        if run_len > worst_run[1]:
            worst_run = (run_start, run_len)

        if worst_run[1] > MAX_CONSECUTIVE:
            start = worst_run[0]
            end   = start + timedelta(days=worst_run[1] - 1)
            plate_violations.append({
                "type":    "consecutive",
                "details": (
                    f"{worst_run[1]} consecutive days "
                    f"({start.strftime('%m/%d/%Y')} – {end.strftime('%m/%d/%Y')})"
                ),
            })

        # ── Check 2: rolling 30-day window (today − 29 days → today) ─────
        days_in_window = [d for d in sorted_dates if window_start <= d <= today]
        if len(days_in_window) > MAX_DAYS_IN_30_DAY:
            plate_violations.append({
                "type":    "30-day",
                "details": (
                    f"{len(days_in_window)} days in the last 30 days "
                    f"({window_start.strftime('%m/%d/%Y')} – {today.strftime('%m/%d/%Y')})"
                ),
            })

        if plate_violations:
            permit_info    = plate_permit.get(plate)
            permit_label   = permit_info[1] if permit_info else "N/A"
            citation_dates = sorted(plate_citations[plate])
            results[plate] = {
                "permit":         permit_label,
                "total_days":     len(sorted_dates),
                "first_seen":     sorted_dates[0],
                "last_seen":      sorted_dates[-1],
                "violations":     plate_violations,
                "citation_count": len(citation_dates),
                "citation_dates": citation_dates,
            }

    # ── Unpermitted citation violations ─────────────────────────────────────
    # Vehicles with no valid permit and >= UNPERMITTED_TOW_CITATIONS WARNINGs
    # in the 30-day window.  Excludes plates already in rule_d results,
    # excluded plates/permits, and plates that have ever had a numeric permit.
    citation_results = {}
    for plate, cite_dates in plate_citations.items():
        if plate in results or plate in EXCLUDED_PLATES:
            continue
        permit_info = plate_permit.get(plate)
        if permit_info and str(permit_info[1]).upper() in EXCLUDED_PERMITS:
            continue
        if plate in plate_has_permit:
            continue  # has a valid parking pass — Rule D applies instead
        sorted_cites = sorted(cite_dates)
        if len(sorted_cites) >= UNPERMITTED_TOW_CITATIONS:
            all_dates = sorted(d for d in plate_dates[plate]
                               if window_start <= d <= today)
            citation_results[plate] = {
                "citation_count": len(sorted_cites),
                "citation_dates": sorted_cites,
                "first_seen":     all_dates[0] if all_dates else sorted_cites[0],
                "last_seen":      all_dates[-1] if all_dates else sorted_cites[-1],
            }

    return results, citation_results


# ── Build email ───────────────────────────────────────────────────────────────
_ROW_COLORS = ["#f9f9f9", "#ffffff"]

def build_email(violations, run_date, citation_violations=None):
    """Returns (subject, html_body). Returns (None, None) if nothing to report."""
    if not violations and not citation_violations:
        return None, None

    citation_violations = citation_violations or {}

    subject = (
        f"Harbor Lights HOA – Guest Parking Violation Report – "
        f"{run_date.strftime('%B %d, %Y')}"
    )

    # Build tow-alert callout for plates at/over the citation threshold
    tow_plates = [
        p for p, info in violations.items()
        if info.get("citation_count", 0) >= CITATION_TOW_THRESHOLD
    ]
    if tow_plates:
        tow_list = "".join(f"<li><strong>{p}</strong></li>" for p in sorted(tow_plates))
        tow_alert_html = f"""
    <div style="background:#fff0f0;border-left:4px solid #b30000;padding:12px 18px;
                margin:0 0 18px;border-radius:0 4px 4px 0;">
      <strong style="color:#b30000;">⚠ Tow Recommendation</strong><br>
      The following vehicle{'s have' if len(tow_plates) != 1 else ' has'} received
      {CITATION_TOW_THRESHOLD} or more citations within the last 30 days and
      {'are' if len(tow_plates) != 1 else 'is'} eligible for immediate tow
      per Rule A (California Vehicle Code Section 22658.2):
      <ul style="margin:8px 0 0;">{tow_list}</ul>
    </div>"""
    else:
        tow_alert_html = ""

    # Build one table row per violation per plate
    rows_html = ""
    for i, (plate, info) in enumerate(sorted(violations.items())):
        row_bg      = _ROW_COLORS[i % 2]
        cite_count  = info.get("citation_count", 0)
        cite_dates  = info.get("citation_dates", [])

        # Citation cell content
        if cite_count == 0:
            cite_html  = '<span style="color:#888;">None on record</span>'
            rec_html   = "—"
            rec_style  = "color:#555;"
        elif cite_count == 1:
            cite_html  = (
                f'<strong style="color:#cc5500;">1 citation</strong><br>'
                f'<span style="font-size:12px;color:#777;">'
                f'({cite_dates[0].strftime("%m/%d/%Y")})</span>'
            )
            rec_html   = "Monitor — 1 more citation triggers tow recommendation"
            rec_style  = "color:#cc5500;"
        elif cite_count >= CITATION_TOW_THRESHOLD:
            date_list  = ", ".join(d.strftime("%m/%d/%Y") for d in cite_dates)
            cite_html  = (
                f'<strong style="color:#b30000;">{cite_count} citations</strong><br>'
                f'<span style="font-size:12px;color:#777;">({date_list})</span>'
            )
            rec_html   = (
                f"⚠ RECOMMEND TOW — {cite_count} citations issued within the last 30 days. "
                f"Vehicle has exhausted the warning process. Eligible for tow per Rule A."
            )
            rec_style  = "color:#b30000;font-weight:bold;"

        for v in info["violations"]:
            vtype_label = (
                "Consecutive Days Exceeded"
                if v["type"] == "consecutive"
                else "30-Day Limit Exceeded"
            )
            vtype_color = "#b30000" if v["type"] == "consecutive" else "#cc5500"
            rows_html += f"""
            <tr style="background:{row_bg};">
              <td style="padding:9px 14px;border:1px solid #ddd;font-weight:bold;">{plate}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;">{info['permit']}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;color:{vtype_color};font-weight:bold;">{vtype_label}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;">{v['details']}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;text-align:center;">{info['total_days']}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;text-align:center;">{info['last_seen'].strftime('%m/%d/%Y')}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;text-align:center;">{cite_html}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;font-size:13px;{rec_style}">{rec_html}</td>
            </tr>"""

    body = f"""
<!DOCTYPE html>
<html>
<body style="font-family:Arial,sans-serif;font-size:14px;color:#333;max-width:900px;margin:0 auto;">

  <div style="background:#1f3864;color:#fff;padding:18px 24px;border-radius:6px 6px 0 0;">
    <h2 style="margin:0;font-size:20px;">Harbor Lights HOA — Guest Parking Violation Report</h2>
    <p style="margin:4px 0 0;font-size:13px;opacity:0.85;">
      Generated {run_date.strftime('%B %d, %Y')} &nbsp;|&nbsp; Americal Patrol Security Services
    </p>
  </div>

  <div style="padding:20px 24px;border:1px solid #ddd;border-top:none;">

    <p>Dear Harbor Lights HOA Board,</p>

    <p>The following is the guest parking violation report for
    <strong>{run_date.strftime('%B %d, %Y')}</strong>.</p>"""

    # ── Rule D section (only if there are Rule D violations) ──────────────
    if violations:
        body += f"""

    <div style="background:#f0f4fa;border-left:4px solid #1f3864;padding:12px 18px;
                margin:16px 0;border-radius:0 4px 4px 0;">
      <strong>Rule D:</strong> &ldquo;Guest,&rdquo; as defined for parking purposes, is a non-resident
      who is invited into Harbor Lights Townhome Association and is making use of Guest Parking for
      no more than <strong>four (4) consecutive days</strong>, or for no more than
      <strong>ten (10) non-consecutive days within a thirty (30) day period.&rdquo;</strong>
    </div>

    <h3 style="color:#b30000;margin-top:24px;">
      Rule D Violations &nbsp;
      <span style="font-size:14px;font-weight:normal;color:#555;">
        ({len(violations)} vehicle{'s' if len(violations) != 1 else ''})
      </span>
    </h3>

    {tow_alert_html}

    <table style="border-collapse:collapse;width:100%;">
      <thead>
        <tr style="background:#1f3864;color:#fff;">
          <th style="padding:10px 14px;text-align:left;">License Plate</th>
          <th style="padding:10px 14px;text-align:left;">Permit #</th>
          <th style="padding:10px 14px;text-align:left;">Violation Type</th>
          <th style="padding:10px 14px;text-align:left;">Details</th>
          <th style="padding:10px 14px;text-align:center;">Total Days<br>on Record</th>
          <th style="padding:10px 14px;text-align:center;">Last Seen</th>
          <th style="padding:10px 14px;text-align:center;">Citations<br>(Last 30 Days)</th>
          <th style="padding:10px 14px;text-align:left;">Recommendation</th>
        </tr>
      </thead>
      <tbody>
        {rows_html}
      </tbody>
    </table>"""

    # ── Unpermitted citation section ─────────────────────────────────────
    if citation_violations:
        cite_rows = ""
        for i, (plate, info) in enumerate(sorted(citation_violations.items())):
            row_bg     = _ROW_COLORS[i % 2]
            cite_count = info["citation_count"]
            date_list  = ", ".join(d.strftime("%m/%d/%Y") for d in info["citation_dates"])
            cite_rows += f"""
            <tr style="background:{row_bg};">
              <td style="padding:9px 14px;border:1px solid #ddd;font-weight:bold;">{plate}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;text-align:center;">
                <strong style="color:#b30000;">{cite_count}</strong>
              </td>
              <td style="padding:9px 14px;border:1px solid #ddd;font-size:13px;">{date_list}</td>
              <td style="padding:9px 14px;border:1px solid #ddd;text-align:center;">{info['last_seen'].strftime('%m/%d/%Y')}</td>
            </tr>"""

        cite_list = "".join(
            f"<li><strong>{p}</strong></li>"
            for p in sorted(citation_violations)
        )
        body += f"""

    <h3 style="color:#b30000;margin-top:30px;">
      Unpermitted Vehicles — Tow Authorization Request &nbsp;
      <span style="font-size:14px;font-weight:normal;color:#555;">
        ({len(citation_violations)} vehicle{'s' if len(citation_violations) != 1 else ''})
      </span>
    </h3>

    <div style="background:#fff0f0;border-left:4px solid #b30000;padding:12px 18px;
                margin:0 0 18px;border-radius:0 4px 4px 0;">
      <strong style="color:#b30000;">⚠ Tow Authorization Requested</strong><br>
      The following vehicle{'s have' if len(citation_violations) != 1 else ' has'}
      been observed parking <strong>without a valid permit</strong> and
      {'have' if len(citation_violations) != 1 else 'has'} received
      {UNPERMITTED_TOW_CITATIONS} or more parking citations within the last 30 days.
      We are requesting authorization to tow on the next offense, per Rule A
      (California Vehicle Code Section 22658.2):
      <ul style="margin:8px 0 0;">{cite_list}</ul>
    </div>

    <table style="border-collapse:collapse;width:100%;">
      <thead>
        <tr style="background:#1f3864;color:#fff;">
          <th style="padding:10px 14px;text-align:left;">License Plate</th>
          <th style="padding:10px 14px;text-align:center;">Citations<br>(Last 30 Days)</th>
          <th style="padding:10px 14px;text-align:left;">Citation Dates</th>
          <th style="padding:10px 14px;text-align:center;">Last Seen</th>
        </tr>
      </thead>
      <tbody>
        {cite_rows}
      </tbody>
    </table>"""

    body += f"""

    <p style="margin-top:24px;color:#555;font-size:13px;">
      Per <strong>Rule A</strong>, vehicles in violation may be towed at the owner&rsquo;s expense
      without warning, in accordance with California Vehicle Code Section 22658.2. Please review
      this report and take appropriate action as needed.
    </p>

    <p style="margin-top:24px;">
      Best Regards,<br>
      <strong>Larry</strong><br><br>
      <strong>Americal Patrol, Inc.</strong><br>
      Mailing: 3301 Harbor Blvd., Oxnard, CA 93035<br>
      VC Office: (805) 844-9433&nbsp;&nbsp;|&nbsp;&nbsp;LA &amp; OC Office: (714) 521-0855&nbsp;&nbsp;|&nbsp;&nbsp;FAX: (866) 526-8472<br>
      <a href="http://www.americalpatrol.com" style="color:#1f3864;">www.americalpatrol.com</a>
    </p>

  </div>

</body>
</html>
"""
    return subject, body


# ── Create Gmail draft ────────────────────────────────────────────────────────
def create_draft(service, subject, body_html, recipients, cc=None):
    msg            = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["To"]      = ", ".join(recipients)
    if cc:
        msg["Cc"]  = ", ".join(cc)
    msg.attach(MIMEText(body_html, "html"))
    raw   = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    draft = service.users().drafts().create(
        userId="me",
        body={"message": {"raw": raw}},
    ).execute()
    return draft["id"]


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    today = date.today()
    print(f"\nHarbor Lights Parking Audit — {today.strftime('%Y-%m-%d')}")
    log.info("=== Parking Audit started ===")

    print("Reading parking data...")
    records = read_parking_records()
    print(f"  {len(records)} records loaded.")

    print("Checking violations...")
    violations, citation_violations = find_violations(records, today)

    # ── Log Rule D violations ─────────────────────────────────────────────
    if violations:
        print(f"  {len(violations)} Rule D violation(s):")
        for plate, info in sorted(violations.items()):
            cite_count = info.get("citation_count", 0)
            for v in info["violations"]:
                line = (
                    f"    {plate} (permit: {info['permit']}): {v['details']}"
                    f"  |  citations (30-day): {cite_count}"
                )
                print(line)
                log.info(line)
            if cite_count >= CITATION_TOW_THRESHOLD:
                tow_msg = f"    *** TOW RECOMMENDED: {plate} has {cite_count} citations in last 30 days ***"
                print(tow_msg)
                log.warning(tow_msg)

    # ── Log unpermitted citation violations ────────────────────────────────
    if citation_violations:
        print(f"  {len(citation_violations)} unpermitted vehicle(s) with 2+ citations:")
        for plate, info in sorted(citation_violations.items()):
            date_list = ", ".join(d.strftime("%m/%d") for d in info["citation_dates"])
            line = (
                f"    {plate}: {info['citation_count']} citations ({date_list})"
                f"  — requesting tow authorization"
            )
            print(line)
            log.info(line)

    if not violations and not citation_violations:
        msg = "No violations found. No email drafted."
        print(f"  {msg}")
        log.info(msg)
        return

    subject, body_html = build_email(violations, today, citation_violations)

    print("\nConnecting to Gmail...")
    service  = get_gmail_service()
    draft_id = create_draft(service, subject, body_html, HOA_RECIPIENTS, cc=CC_LIST)

    print(f"  Draft created  : {draft_id}")
    print(f"  Subject        : {subject}")
    print(f"  To             : {', '.join(HOA_RECIPIENTS)}")
    print(f"  CC             : {', '.join(CC_LIST)}")
    log.info("Draft created: %s | Subject: %s", draft_id, subject)

    print("\nDone.")
    log.info("=== Parking Audit complete ===")


if __name__ == "__main__":
    main()
